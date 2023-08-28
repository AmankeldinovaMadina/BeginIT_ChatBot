import telebot
from telebot import types
import datetime as dt
from openpyxl import load_workbook
import schedule
import time
from datetime import datetime

bot = telebot.TeleBot('6584233589:AAHMVktC7Ji51ZN_LHqsjnSIss9SK9H2LcE')
message_admin = {}
workbook = load_workbook('ChatBot.xlsx')
worksheet = workbook['Sheet1']
count = worksheet.max_row + 1
group = " "
groupA = 0
groupB = 0
groupC = 0
groupD = 0
chatID = None
category = " "
@bot.message_handler(commands=["start"])
def start(message):
    global count
    chat_id = message.chat.id
    worksheet[f'A{count}'].value = str(dt.datetime.now().date())
    worksheet[f'B{count}'].value = str(dt.datetime.now().time())[0:8]
    worksheet[f'C{count}'].value = f"{message.chat.first_name} {message.chat.last_name}"
    worksheet[f'D{count}'].value = message.chat.username
    
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Тіркелу/Зарегистрироваться'))
    chat_id = message.chat.id
    first_name = message.chat.first_name
    bot.send_sticker(chat_id, sticker=r'CAACAgIAAxkBAAEImRFkO3MzP7AWT8T1uuATaNRRLJav4gACHwADWbv8Jeo5dBvZPTaZLwQ')
    bot.send_message(chat_id, f"Cалем {first_name}!\n"
                     f"BeginIT-ге қош келдің!\n\n\nПривет {first_name}.\nДобро пожаловать на BeginIT! ", reply_markup=app_markup)
    pdf_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Инструкция_по_использованию_Notion👋.pdf?alt=media&token=caca0003-57fc-4d3b-ab81-434a44cfa1e1'
    bot.send_document(chat_id, pdf_url, caption='Алдағы уақытта барлық тапсырмаларды notion-да жасайтын боламыз. Мына pdf ішінен notion-ды қалай қолдану керектігі жайында барлық ақпаратты таба аласың. \n\n\nВ дальнейшем мы будем выполнять все поставленные задачи в notion. Всю информацию о том, как использовать понятие, вы можете найти в этом PDF-файле. \n\n\nАл бұл notion ссылкасы: https://sleepy-countess-e1f.notion.site/BeginIT-Summer-Fest-60ee365a535c4b338139e03c030936fc \n\n\n А это ссылка на notion: https://sleepy-countess-e1f.notion.site/BeginIT-Summer-Fest-60ee365a535c4b338139e03c030936fc')
    

@bot.message_handler(content_types=["text"])
def text(message):
    global chatID
    chat_id = message.chat.id
    chatID = chat_id
    if message.text == 'Тіркелу/Зарегистрироваться':
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Мұғалім/Учитель'))
        app_markup.add(types.KeyboardButton('Оқушы/Ученик'))
        app_markup.add(types.KeyboardButton('Спикер'))

        bot.send_message(chat_id, "Өз категорияңызды таңдаңыз. \n\nВыберите свою категорию.", reply_markup=app_markup)
          
    if category == 'Спикер':
       bot.register_next_step_handler(message, speaker_choice)  
    else:
        bot.register_next_step_handler(message, select_level)

    
def send_dev(message):
    chat_id = message.chat.id
    bot.send_message(f"Поступило сообщение о работе бота!\n\n"
                     f"Текст: {message.text}\n\n"
                     f"Пользователь: {message.chat.first_name}\n"
                     f"Username: @{message.chat.username}")
    bot.send_message(chat_id, f"Ваше сообщение отправлено разработчикам!\nСпасибо за обратную связь 🙏\n\nДля повторного запуска напишите /start")


@bot.message_handler(content_types=["text"])
def select_level(message):
    global count, groupA, groupB, groupC, groupD, group, category 
    worksheet[f"E{count}"].value = message.text
    message_admin['who'] = message.text
    
    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Айтыскерлер'), types.KeyboardButton('Қолөнершілер'),  types.KeyboardButton('Саудагерлер'),  types.KeyboardButton('Саяхатшылар'))
    bot.send_message(chat_id, "Өз тобыңызды таңдаңыз:\n\nВыберите свою группу:", reply_markup=app_markup)
    message_admin['level'] = message.text
    category = message.text

    if category == 'Спикер':
       bot.register_next_step_handler(message, speaker_choice)
    else:
       bot.register_next_step_handler(message, select_info)

@bot.message_handler(content_types=["text"])
def speaker_choice(message):
    global count, groupA, groupB, groupC, groupD, group, category 
    
    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Айтыскерлер'), types.KeyboardButton('Қолөнершілер'),  types.KeyboardButton('Саудагерлер'),  types.KeyboardButton('Саяхатшылар'))
    bot.send_message(chat_id, "Қай топқа балл беріледі?:\n\nКакая группа получает балл?:", reply_markup=app_markup)
    bot.register_next_step_handler(message, speaker_pointAdding)

@bot.message_handler(content_types=["text"])
def speaker_pointAdding(message):
    global count, groupA, groupB, groupC, groupD, group, category 
    
    chat_id = message.chat.id
    chosenGroup = message.text
    if chosenGroup == 'Айтыскерлер':
            groupA = groupA + 1
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, speaker_choice)
    elif chosenGroup == 'Қолөнершілер':
            groupA = groupB + 1
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, speaker_choice)
    elif chosenGroup == 'Саудагерлер':
            groupA = groupC + 1
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, speaker_choice)
    elif chosenGroup == 'Саяхатшылар':
            groupA = groupD + 1
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, speaker_choice)



@bot.message_handler(content_types=["text"])
def send_info_AstanaHub():
    global chatID
    chat_id = chatID
    bot.send_message(chat_id, "Астана Hub Это международный технопарк IT-стартапов. Здесь создаются условия для свободного развития казахстанских и зарубежных технологических компаний. Astana Hub стремится стать центром развития инновационных проектов, выпускать прорывные IT-компании, а также стать очагом для притяжения критической массы молодых и талантливых IT-специалистов со всего мира. Деятельность офиса образовательных программ Astana Hub нацелена на подпитывание пула казахстанских талантов в сфере IT через популяризацию технологического предпринимательства и обучение навыкам, необходимым для создания технологического стартапа.\nАстана Hub  бұл IT-стартаптардың халықаралық технопаркі. Мұнда қазақстандық және шетелдік технологиялық компаниялардың еркін дамуына жағдай жасалынады.\nAstana Hub миссиясы – инновациялық жобаларды дамыту орталығына айналу, серпінді IT-компанияларды шығару және әлемнің түкпір-түкпірінен жас және дарынды IT мамандарының сыни массасын тартудың ошағына айналу.\nAstana Hub білім беру бағдарламалары кеңсесінің қызметі технологиялық кәсіпкерлікті танымал ету және технологиялық стартап құру үшін қажетті дағдыларды оқыту арқылы IT саласындағы қазақстандық таланттар пулын арттыруға бағытталған.")

schedule.every().day.at("09:45").do(send_info_AstanaHub)

@bot.message_handler(content_types=["text"])
def send_info_AstanaIT():
    global chatID
    chat_id = chatID
    bot.send_message(chat_id, "Астана IT университет \nAstana IT University – лидер инновационного непрерывного IT-образования и науки с устойчивыми академическими традициями и высокой социальной ответственностью \nГлобальная цель - Обеспечение качества подготовки кадров в сфере ИКТ на междисциплинарной основе \n\n\nAstana IT University-тұрақты академиялық дәстүрлері мен жоғары әлеуметтік жауапкершілігі бар инновациялық үздіксіз IT-білім беру мен ғылымның көшбасшысы \nУниверситет мақсаты - Пәнаралық негізде АКТ саласында кадрларды даярлау сапасын қамтамасыз ету")

schedule.every().day.at("15:55").do(send_info_AstanaIT)

@bot.message_handler(content_types=["text"])
def send_info_Bouling():
    global chatID
    chat_id = chatID
    bot.send_message(chat_id, "Бүгін өте нәтижелі күн болды, енді көңіл көтеруге баратын кез келді!  Боулингке барамыз, егер қалай ойнайтынын білмесеңіз -бірге үйренеміз!")

schedule.every().day.at("17:35").do(send_info_Bouling)

@bot.message_handler(content_types=["text"])
def send_info_SpeakerOne():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1zIgFOhUyaUrrvfvDxuSj4TZT2Ecii6FQ/view'
    bot.send_photo(chat_id, photo_url, caption='Baglan Kutubayev \nEdTech startup entrepreneur, Artisan Education, \nEx-Vice Principal, Quantum STEM School \n\nТема выступления:Information Technologies: Yesterday, Today, Tomorrow')

schedule.every().day.at("10:35").do(send_info_SpeakerOne)

@bot.message_handler(content_types=["text"])
def send_info_SpeakerTwo():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1wqnPX48Sz_gddfwR4x9ZGRvs465_dhSm/view'
    bot.send_photo(chat_id, photo_url, caption='Alikhan Talipov \nBeginIT tracker-teacher, EPAM devrel \n\nDirections and Professions in IT')

schedule.every().day.at("11:10").do(send_info_SpeakerTwo)

@bot.message_handler(content_types=["text"])
def send_info_SpeakerThree():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1krZUNFfH2SuILLoE6oENePvf2gcG05Te/view?usp=sharing'
    bot.send_photo(chat_id, photo_url, caption='Daulet Beimurzinov  \nDirector of Education Programs Office of Astana HUB \n\nEntry Points into IT')

schedule.every().day.at("12:00").do(send_info_SpeakerThree)



@bot.message_handler(content_types=["text"])
def send_info_SpeakerFour():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1sSr_Qz2zVJcyFSVq_NyYXlm1KyTZlIvG/view?usp=sharing '
    bot.send_photo(chat_id, photo_url, caption='Vitaliy Perov  \nTSARKA expert on Offensive security and cyber intelligence, Digital Forensic. \n\nOSINT basics in practice')

schedule.every().day.at("12:35").do(send_info_SpeakerFour)

@bot.message_handler(content_types=["text"])
def send_info_SpeakerFive():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1yxW4PugyaYS0laquNKAWO6dMGDSfEv9J/view?usp=sharing '
    bot.send_photo(chat_id, photo_url, caption='Yersain Kabdrashev  \nSenior educational project specialist  \n\nWho Makes the inDrive App?')

schedule.every().day.at("14:58").do(send_info_SpeakerFive)

@bot.message_handler(content_types=["text"])
def send_info_SpeakerMadina():
    global chatID
    chat_id = chatID
    photo_url = 'https://drive.google.com/file/d/1yxW4PugyaYS0laquNKAWO6dMGDSfEv9J/view?usp=sharing '
    bot.send_photo(chat_id, photo_url, caption='Yersain Kabdrashev  \nSenior educational project specialist  \n\nWho Makes the inDrive App?')

schedule.every().day.at("07:10").do(send_info_SpeakerMadina)

@bot.message_handler(content_types=["text"])
def select_info(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f'F{count}'].value = message.text
    message_admin['level'] = message.text
    

    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Help'))
    bot.send_message(chat_id, "Тапсырма 1.«Қазына іздеу»\nТапсырма: орынды зерттеңіз, технологиялық жаңалықтарды табыңыз. Notion-да фото коллаж жасаңыз. \nБағалау: коллаж жібергеніңіз үшін 7 қошқар аласыз\n\n\nЗадание 1. «Охота за сокровищами»\nЗадание: Исследуйте место, найдите технологические новшества. Создайте в notion фото-коллаж. \nОценка: за отправку коллажа получаете 7 кошкарчиков", reply_markup=app_markup)
    bot.register_next_step_handler(message, select_info_yn)
    group = message.text
    print(group)
    message_admin['info'] = message.text

@bot.message_handler(content_types=["document", "text"])
def select_info_yn(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"G{count}"].value = message.text
    message_admin['info'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Сұрақтарға жауап беріңіз: өзіңіз үшін не жаңалық көрдіңіз?Не таң қалдырды? Сіз өзіңіз үшін қандай қорытынды жасадыңыз? Бұл сіздің болашақта мамандық таңдауыңызға қалай әсер етуі мүмкін?\n\n\nОтветьте на вопросы: Что нового вы увидели для себя? Что удивило? Какие выводы вы сделали для себя? Как это может повлиять на ваш выбор профессии в будущем?", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Қолөнершілер':
            groupB = groupB + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Саудагерлер':
            groupC = groupC + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Саяхатшылар':
            groupD = groupD + 7
            bot.register_next_step_handler(message, event_date)
    else:
        if group == 'Айтыскерлер':
            groupA = groupA + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Қолөнершілер':
            groupA = groupB + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Саудагерлер':
            groupA = groupC + 7
            bot.register_next_step_handler(message, event_date)
        elif group == 'Саяхатшылар':
            groupA = groupD + 7
            bot.register_next_step_handler(message, event_date)

@bot.message_handler(content_types=["text"])
def event_date(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"H{count}"].value = message.text
    message_admin['short_about_event'] = message.text
    
    chat_id = message.chat.id
    photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
    bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')


    bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
    bot.send_message(chat_id, "Тапсырма 2. \nДәріс \nАқпараттық технологиялар: кеше, бүгін, ертең \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\nЗадание 2. \nИнформационные технологии: вчера, сегодня, завтра \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала.\nОтветьте устно.\nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
    bot.register_next_step_handler(message, event_question)

@bot.message_handler(content_types=["text"])
def event_question(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"I{count}"].value = message.text
    message_admin['short_about_event'] = message.text
    
    chat_id = message.chat.id
    if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "3-тапсырма. \nДәріс \nIT бағыттары мен мамандықтары\nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\nЗадание 3. \nЛекцияНаправления и профессии в IT \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionTwo)
    elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "3-тапсырма. \nДәріс \nIT бағыттары мен мамандықтары\nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\nЗадание 3. \nЛекцияНаправления и профессии в IT \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionTwo)
    elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "3-тапсырма. \nДәріс \nIT бағыттары мен мамандықтары\nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\n3-тапсырма. \nДәріс \nIT бағыттары мен мамандықтары\nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\nЗадание 3. \nЛекцияНаправления и профессии в IT \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionTwo)
    elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "3-тапсырма. \nДәріс \nIT бағыттары мен мамандықтары\nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.\nАуызша жауап беріңіз.\nБағалау: Жауап үшін 1 қошқар\n\n\nЗадание 3. \nЛекцияНаправления и профессии в IT \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionTwo)

@bot.message_handler(content_types=["text"])
def event_questionTwo(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"G{count}"].value = message.text
    message_admin['short_about_event'] = message.text
    
    chat_id = message.chat.id
    if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "4 тапсырма. \nДәріс \nIT жолын неден бастау керек және IT-ге кіру нүктелері қандай? \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.Ауызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 4. \nЛекция \nС чего начать путь в IT и какие есть точки входа в IT? \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionThree)
    elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "4 тапсырма. \nДәріс IT жолын неден бастау керек және IT-ге кіру нүктелері қандай? \nдәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.ауызша жауап беріңіз. \nбағалау: жауап үшін 1 қошқар \n\n\nзадание 4. \nлекция с чего начать путь в it и какие есть точки входа в it? \nлектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nответьте устно. \nоценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionThree)
    elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "4 тапсырма. \nДәріс IT жолын неден бастау керек және IT-ге кіру нүктелері қандай? \nдәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.ауызша жауап беріңіз. \nбағалау: жауап үшін 1 қошқар \n\n\nзадание 4. \nлекция с чего начать путь в it и какие есть точки входа в it? \nлектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nответьте устно. \nоценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionThree)
    elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "4 тапсырма. \nДәріс IT жолын неден бастау керек және IT-ге кіру нүктелері қандай? \nдәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды.ауызша жауап беріңіз. \nбағалау: жауап үшін 1 қошқар \n\n\nзадание 4. \nлекция с чего начать путь в it и какие есть точки входа в it? \nлектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nответьте устно. \nоценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionThree)


    

@bot.message_handler(content_types=["text"])
def event_questionThree(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"K{count}"].value = message.text
    message_admin['short_about_event'] = message.text
    
    chat_id = message.chat.id
    if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_message(chat_id, "5 тапсырма.\n Дәріс \nТәжірибедегі OSINT негіздері \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 5. \nЛекцияОсновы OSINT на практике \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionFour)
    elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "5 тапсырма.\n Дәріс \nТәжірибедегі OSINT негіздері \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 5. \nЛекцияОсновы OSINT на практике \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionFour)
    elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "5 тапсырма.\n Дәріс \nТәжірибедегі OSINT негіздері \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 5. \nЛекцияОсновы OSINT на практике \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionFour)
    elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "5 тапсырма.\n Дәріс \nТәжірибедегі OSINT негіздері \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 5. \nЛекцияОсновы OSINT на практике \nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_questionFour)




@bot.message_handler(content_types=["text"])
def event_questionFour(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"L{count}"].value = message.text
    message_admin['date'] = message.text
    
    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Help'))
    if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "6 тапсырма. \nIT саласы және болашақ мамандықтары туралы сұхбат жүргізу \nАлақай! «Әлемді құтқаруға көмектесетін болашақтың 5 IT мамандығы» атты шағын зерттеуді бастау уақыты келді.    \nБағалау: Чат-бот арқылы әрбір рәсімделіп жіберілген жауап үшін 1 қошқар \n\n\nЗадание 6. Провести интервью о сфере IT и профессиях будущего \nУра! Пора начать мини-исследование «5 IT профессий будущего, которые помогут спасти мир». \nОценка по 1 кошкарчику каждому приславшему в оформленный ответ через чат-бот", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_size)
    elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "6 тапсырма. \nIT саласы және болашақ мамандықтары туралы сұхбат жүргізу \nАлақай! «Әлемді құтқаруға көмектесетін болашақтың 5 IT мамандығы» атты шағын зерттеуді бастау уақыты келді.    \nБағалау: Чат-бот арқылы әрбір рәсімделіп жіберілген жауап үшін 1 қошқар \n\n\nЗадание 6. Провести интервью о сфере IT и профессиях будущего \nУра! Пора начать мини-исследование «5 IT профессий будущего, которые помогут спасти мир». \nОценка по 1 кошкарчику каждому приславшему в оформленный ответ через чат-бот", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_size)
    elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "6 тапсырма. \nIT саласы және болашақ мамандықтары туралы сұхбат жүргізу \nАлақай! «Әлемді құтқаруға көмектесетін болашақтың 5 IT мамандығы» атты шағын зерттеуді бастау уақыты келді.    \nБағалау: Чат-бот арқылы әрбір рәсімделіп жіберілген жауап үшін 1 қошқар \n\n\nЗадание 6. Провести интервью о сфере IT и профессиях будущего \nУра! Пора начать мини-исследование «5 IT профессий будущего, которые помогут спасти мир». \nОценка по 1 кошкарчику каждому приславшему в оформленный ответ через чат-бот", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_size)
    elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "6 тапсырма. \nIT саласы және болашақ мамандықтары туралы сұхбат жүргізу \nАлақай! «Әлемді құтқаруға көмектесетін болашақтың 5 IT мамандығы» атты шағын зерттеуді бастау уақыты келді.    \nБағалау: Чат-бот арқылы әрбір рәсімделіп жіберілген жауап үшін 1 қошқар \n\n\nЗадание 6. Провести интервью о сфере IT и профессиях будущего \nУра! Пора начать мини-исследование «5 IT профессий будущего, которые помогут спасти мир». \nОценка по 1 кошкарчику каждому приславшему в оформленный ответ через чат-бот", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_size)


@bot.message_handler(content_types=["text"])
def event_size(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"M{count}"].value = message.text
    message_admin['link'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Спроси мнения экспертов, участников, собравшихся. Например, можно спросить: Как думаете, насколько информационные технологии будут играть роль в решении глобальных проблем? Какие технологии будут самыми востребованными через 30 лет?Как бы вы назвали эти профессии?Результаты всех наблюдений и интервью напишите, оформите в  notion и загрузите ответ через чат-бот.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
    else:
        if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)
        elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_count)



@bot.message_handler(content_types=["text"])
def event_count(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"N{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Пропустить задание'))
    app_markup.add(types.KeyboardButton('Help'))
    if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "7 тапсырма. \nПитчинг аймағы \nBeginITbyinDrive - те оқыған No code және Digital design бағдарламалары есіңде бар ма?   \nТапсырма өз қалауың бойынша орындалады.  \nӨзіңнің жобаңды ұсынып, 15 қошқар таба аласың. \n\n\nЗадание 7.\nПитчинг-зона \nА ты помнишь про программы No code и Digital design, которые изучали в BeginIT by inDrive? \nЗадание выполняется по желанию. \nМожешь запитчить свой проект, и заработать 15 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_taskEight)
    elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "7 тапсырма. \nПитчинг аймағы \nBeginITbyinDrive - те оқыған No code және Digital design бағдарламалары есіңде бар ма?   \nТапсырма өз қалауың бойынша орындалады.  \nӨзіңнің жобаңды ұсынып, 15 қошқар таба аласың. \n\n\nЗадание 7.\nПитчинг-зона \nА ты помнишь про программы No code и Digital design, которые изучали в BeginIT by inDrive? \nЗадание выполняется по желанию. \nМожешь запитчить свой проект, и заработать 15 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_taskEight)
    elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "7 тапсырма. \nПитчинг аймағы \nBeginITbyinDrive - те оқыған No code және Digital design бағдарламалары есіңде бар ма?   \nТапсырма өз қалауың бойынша орындалады.  \nӨзіңнің жобаңды ұсынып, 15 қошқар таба аласың. \n\n\nЗадание 7.\nПитчинг-зона \nА ты помнишь про программы No code и Digital design, которые изучали в BeginIT by inDrive? \nЗадание выполняется по желанию. \nМожешь запитчить свой проект, и заработать 15 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_taskEight)
    elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "7 тапсырма. \nПитчинг аймағы \nBeginITbyinDrive - те оқыған No code және Digital design бағдарламалары есіңде бар ма?   \nТапсырма өз қалауың бойынша орындалады.  \nӨзіңнің жобаңды ұсынып, 15 қошқар таба аласың. \n\n\nЗадание 7.\nПитчинг-зона \nА ты помнишь про программы No code и Digital design, которые изучали в BeginIT by inDrive? \nЗадание выполняется по желанию. \nМожешь запитчить свой проект, и заработать 15 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_taskEight)


@bot.message_handler(content_types=["text"])
def event_taskEight(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"O{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Можно рассказать о своем проекте/идее собравшимся, например экспертам, в формате Elevator Pitch. Можно запитчить свой проект, который вы делали на программе BeginIT by inDrive. Elevator Pitch - короткий рассказ о концепции продукта, проекта или сервиса. Термин отражает ограниченность по времени - длина презентации должна быть такой, чтобы она могла быть полностью рассказана за время поездки на лифте, то есть от 30 секунд до 1-2 минут.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 15
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Қолөнершілер':
            groupB = groupB + 15
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Саудагерлер':
            groupC = groupC + 15
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Саяхатшылар':
            groupD = groupD + 15
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskNine)

    else:
        if group == 'Айтыскерлер':
            groupA = groupA + 15
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            app_markup.add(types.KeyboardButton('Help'))
            bot.send_message(chat_id, "8 тапсырма. \nWho makes inDrive app? \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 8. \nWho makes inDrive app?\nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Қолөнершілер':
            groupB = groupB + 15
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            app_markup.add(types.KeyboardButton('Help'))
            bot.send_message(chat_id, "8 тапсырма. \nWho makes inDrive app? \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 8. \nWho makes inDrive app?\nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Саудагерлер':
            groupC = groupC + 15
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            app_markup.add(types.KeyboardButton('Help'))
            bot.send_message(chat_id, "8 тапсырма. \nWho makes inDrive app? \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 8. \nWho makes inDrive app?\nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskNine)
        elif group == 'Саяхатшылар':
            groupD = groupD + 15
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 15 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            app_markup.add(types.KeyboardButton('Help'))
            bot.send_message(chat_id, "8 тапсырма. \nWho makes inDrive app? \nДәріскер дәріс соңында материалды бекітуге арналған 5 сұрақ қояды. \nАуызша жауап беріңіз. \nБағалау: Жауап үшін 1 қошқар \n\n\nЗадание 8. \nWho makes inDrive app?\nЛектор озвучивает 5 вопросов в конце лекции на закрепление материала. \nОтветьте устно. \nОценка 1 кошкарчик за ответ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskNine)



@bot.message_handler(content_types=["text"])
def event_TaskNine(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"P{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Оценивается скорость и правильность ответа. В каждом вопросе 1 победитель.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskTen)
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        if group == 'Айтыскерлер':
            groupA = groupA + 0
            print(groupA)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "9 тапсырма. \nБолашақта IT мамандықтарын қалай көремін \nБолашақта IT мамандықтарын қалай көретініңді айтып бер?  Кем дегенде 1 мамандықты таңдап, топта талқылаңыздар және берілген реквизиттерді қолданыңыздар.  \nБолашақта IT мамандығын сипаттайтын кілт сөздердің тізімін (кем дегенде 10) жазыңыз .  Сөздердің тізімін чат-бот арқылы жіберу керек.\nТопта талқылау үшін 7 минут беріледі.   \nОрындалу уақыты:  \nӘр топтың сөйлеуіне 1 минут уақыт беріледі. \nТапсырманы орындау үшін әр топқа 10 қошқар беріледі. \n\n\nЗадание 9. \nКак я вижу IT -профессии в будущем \nПоделись как ты видишь IT -профессии в будущем? Выбери не менее 1 профессии, обсудите в команде и задействуйте предоставленный реквизит.\nВыпишите список ключевых слов (не менее 10), характеризующую видение IT профессии в будущем. Список слов отправить через чат-бот. \nОбсуждение в командах- 7 минут. \nВремя выполнения: \nВыступление каждой команды – по 1 минуте.\nЗа выполнение задания по 10 кошкарчиков каждой команде", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Қолөнершілер':
            groupB = groupB + 0
            print(groupB)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "9 тапсырма. \nБолашақта IT мамандықтарын қалай көремін \nБолашақта IT мамандықтарын қалай көретініңді айтып бер?  Кем дегенде 1 мамандықты таңдап, топта талқылаңыздар және берілген реквизиттерді қолданыңыздар.  \nБолашақта IT мамандығын сипаттайтын кілт сөздердің тізімін (кем дегенде 10) жазыңыз .  Сөздердің тізімін чат-бот арқылы жіберу керек. \nТопта талқылау үшін 7 минут беріледі.   \nОрындалу уақыты:  \nӘр топтың сөйлеуіне 1 минут уақыт беріледі. \nТапсырманы орындау үшін әр топқа 10 қошқар беріледі. \n\n\nЗадание 9. \nКак я вижу IT -профессии в будущем \nПоделись как ты видишь IT -профессии в будущем? Выбери не менее 1 профессии, обсудите в команде и задействуйте предоставленный реквизит.\nВыпишите список ключевых слов (не менее 10), характеризующую видение IT профессии в будущем. Список слов отправить через чат-бот. \nОбсуждение в командах- 7 минут. \nВремя выполнения: \nВыступление каждой команды – по 1 минуте.\nЗа выполнение задания по 10 кошкарчиков каждой команде", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Саудагерлер':
            groupC = groupC + 0
            print(groupC)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "9 тапсырма. \nБолашақта IT мамандықтарын қалай көремін \nБолашақта IT мамандықтарын қалай көретініңді айтып бер?  Кем дегенде 1 мамандықты таңдап, топта талқылаңыздар және берілген реквизиттерді қолданыңыздар.  \nБолашақта IT мамандығын сипаттайтын кілт сөздердің тізімін (кем дегенде 10) жазыңыз .  Сөздердің тізімін чат-бот арқылы жіберу керек.  \nТопта талқылау үшін 7 минут беріледі.   \nОрындалу уақыты:  \nӘр топтың сөйлеуіне 1 минут уақыт беріледі. \nТапсырманы орындау үшін әр топқа 10 қошқар беріледі. \n\n\nЗадание 9. \nКак я вижу IT -профессии в будущем \nПоделись как ты видишь IT -профессии в будущем? Выбери не менее 1 профессии, обсудите в команде и задействуйте предоставленный реквизит.\nВыпишите список ключевых слов (не менее 10), характеризующую видение IT профессии в будущем. Список слов отправить через чат-бот. \nОбсуждение в командах- 7 минут. \nВремя выполнения: \nВыступление каждой команды – по 1 минуте.\nЗа выполнение задания по 10 кошкарчиков каждой команде", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTen)
        elif group == 'Саяхатшылар':
            groupD = groupD + 0
            print(groupD)
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "9 тапсырма. \nБолашақта IT мамандықтарын қалай көремін \nБолашақта IT мамандықтарын қалай көретініңді айтып бер?  Кем дегенде 1 мамандықты таңдап, топта талқылаңыздар және берілген реквизиттерді қолданыңыздар.  \nБолашақта IT мамандығын сипаттайтын кілт сөздердің тізімін (кем дегенде 10) жазыңыз .  Сөздердің тізімін чат-бот арқылы жіберу керек.  \nТопта талқылау үшін 7 минут беріледі.   \nОрындалу уақыты:  \nӘр топтың сөйлеуіне 1 минут уақыт беріледі. \nТапсырманы орындау үшін әр топқа 10 қошқар беріледі. \n\n\nЗадание 9. \nКак я вижу IT -профессии в будущем \nПоделись как ты видишь IT -профессии в будущем? Выбери не менее 1 профессии, обсудите в команде и задействуйте предоставленный реквизит.\nВыпишите список ключевых слов (не менее 10), характеризующую видение IT профессии в будущем. Список слов отправить через чат-бот. \nОбсуждение в командах- 7 минут. \nВремя выполнения: Выступление каждой команды – по 1 минуте.\nЗа выполнение задания по 10 кошкарчиков каждой команде", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTen)




@bot.message_handler(content_types=["text"])
def event_TaskTen(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"Q{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "В выполнении задания могут помочь вопросы: Кто как выглядит в профессии? Какая профессия, портрет и почему? Какие у них есть навыки? Какие в 2050 году будут профессии? Как будет выглядеть айтишник, что у него будет в руках? \nВ команде выполняют проект распределившись по ролям. В каждой команде нужно выбрать 2 участников, чтобы представить всем собравшимся карту профессий и список ключевых слов (не менее 10 слов). Использовать информацию проведенного ранее сбора информации по мини-исследованию.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 10
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Қолөнершілер':
            groupB = groupB + 10
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Саудагерлер':
            groupC = groupC + 10
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_TaskEleven)
    else:
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "10 тапсырма. \nҚайда барамыз деп ойлайсың?\nҚалай ойлайсың, біз қайда бара жатырмыз?    \nБірінші дұрыс жауап берушіге қосымша 5 қошқар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "10 тапсырма. \nҚайда барамыз деп ойлайсың?\nҚалай ойлайсың, біз қайда бара жатырмыз?    \nБірінші дұрыс жауап берушіге қосымша 5 қошқар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "10 тапсырма. \nҚайда барамыз деп ойлайсың?\nҚалай ойлайсың, біз қайда бара жатырмыз?    \nБірінші дұрыс жауап берушіге қосымша 5 қошқар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskEleven)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 10 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "10 тапсырма. \nҚайда барамыз деп ойлайсың?\nҚалай ойлайсың, біз қайда бара жатырмыз?    \nБірінші дұрыс жауап берушіге қосымша 5 қошқар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskEleven)
    
@bot.message_handler(content_types=["text"])
def event_TaskEleven(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"R{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    app_markup.add(types.KeyboardButton('Help'))
    if group == 'Айтыскерлер':
            groupA = groupA + 5
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 5 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "11 тапсырма. \nAstana IT university-ге экскурсия \nБіз шағын зерттеуді жалғастырамыз.\nСізге IT-білім беруде не ұнайды?\nСіз көрген зертханалардың қайсысы сізге ұнады және неге? \nБағалау: Әрбір рәсімделген жауапты жіберушіге 1 қошқар \n\n\nЗадание 11. \nЭкскурсия в Astana IT university \nПродолжаем мини-исследование.\nЧто вам нравится больше всего в IT-образовании? \nКакая из увиденных лабораторий вам понравилась и почему?\nОценка по 1 кошкарчику каждому приславшему оформленный ответ ар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskEleven)
    elif group == 'Қолөнершілер':
            groupB = groupB + 5
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 5 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "11 тапсырма. \nAstana IT university-ге экскурсия \nБіз шағын зерттеуді жалғастырамыз.\nСізге IT-білім беруде не ұнайды?\nСіз көрген зертханалардың қайсысы сізге ұнады және неге? \nБағалау: Әрбір рәсімделген жауапты жіберушіге 1 қошқар \n\n\nЗадание 11. \nЭкскурсия в Astana IT university \nПродолжаем мини-исследование.\nЧто вам нравится больше всего в IT-образовании? \nКакая из увиденных лабораторий вам понравилась и почему?\nОценка по 1 кошкарчику каждому приславшему оформленный ответ ар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTwelve)
    elif group == 'Саудагерлер':
            groupC = groupC + 5
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 5 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "11 тапсырма. \nAstana IT university-ге экскурсия \nБіз шағын зерттеуді жалғастырамыз.\nСізге IT-білім беруде не ұнайды?\nСіз көрген зертханалардың қайсысы сізге ұнады және неге? \nБағалау: Әрбір рәсімделген жауапты жіберушіге 1 қошқар \n\n\nЗадание 11. \nЭкскурсия в Astana IT university \nПродолжаем мини-исследование.\nЧто вам нравится больше всего в IT-образовании? \nКакая из увиденных лабораторий вам понравилась и почему?\nОценка по 1 кошкарчику каждому приславшему оформленный ответ ар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTwelve)
    elif group == 'Саяхатшылар':
            groupD = groupD + 5
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 5 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.send_message(chat_id, "11 тапсырма. \nAstana IT university-ге экскурсия \nБіз шағын зерттеуді жалғастырамыз.\nСізге IT-білім беруде не ұнайды?\nСіз көрген зертханалардың қайсысы сізге ұнады және неге? \nБағалау: Әрбір рәсімделген жауапты жіберушіге 1 қошқар \n\n\nЗадание 11. \nЭкскурсия в Astana IT university \nПродолжаем мини-исследование.\nЧто вам нравится больше всего в IT-образовании? \nКакая из увиденных лабораторий вам понравилась и почему?\nОценка по 1 кошкарчику каждому приславшему оформленный ответ ар \n\n\nЗадание 10. \nУгадай куда идем? \nКак ты думаешь, куда мы идем дальше? \nПервому давшему правильный ответ, дополнительно 5 кошкарчиков ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_TaskTwelve)



@bot.message_handler(content_types=["text"])
def event_TaskTwelve(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"S{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Делайте фото, задавайте вопросы, будьте любознательными! \nНапример, спросите:В чем принципиальные различия деятельности Лабораторий KUKA, Робототехники и Индустрии 4.0 ?Результаты всех наблюдений и интервью оформите в notion.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "Айтыскерлер:" + str(groupA) + "\nҚолөнершілер:" + str(groupB) + "\nСаудагерлер:" + str(groupC) + "\nСаяхатшылар:" + str(groupD))
            bot.register_next_step_handler(message, event_Task13)
       
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "12 тапсырма. \nПікірталас \nОяндыңдар ма?   Шағын зерттеуді жалғастыру уақыты келді  \n«Технологиялар болашақ әлемін қалай құтқара алады?» деген пікірталасқа қатыс \nБағалау: әрқайсыңа 1 қошқардан  \n\n\nЗадание 12. \nДискуссия \nНадеемся ты проснулся! Пора продолжить мини-исследование \nУчаствуй в дискуссии «Как технологии могут спасти мир будущего?» \nОценка по 1 кошкарчику каждому ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "12 тапсырма. \nПікірталас \nОяндыңдар ма?   Шағын зерттеуді жалғастыру уақыты келді  \n«Технологиялар болашақ әлемін қалай құтқара алады?» деген пікірталасқа қатыс \nБағалау: әрқайсыңа 1 қошқардан  \n\n\nЗадание 12. \nДискуссия \nНадеемся ты проснулся! Пора продолжить мини-исследование \nУчаствуй в дискуссии «Как технологии могут спасти мир будущего?» \nОценка по 1 кошкарчику каждому ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "12 тапсырма. \nПікірталас \nОяндыңдар ма?   Шағын зерттеуді жалғастыру уақыты келді  \n«Технологиялар болашақ әлемін қалай құтқара алады?» деген пікірталасқа қатыс \nБағалау: әрқайсыңа 1 қошқардан  \n\n\nЗадание 12. \nДискуссия \nНадеемся ты проснулся! Пора продолжить мини-исследование \nУчаствуй в дискуссии «Как технологии могут спасти мир будущего?» \nОценка по 1 кошкарчику каждому", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task13)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "12 тапсырма. \nПікірталас \nОяндыңдар ма?   Шағын зерттеуді жалғастыру уақыты келді  \n«Технологиялар болашақ әлемін қалай құтқара алады?» деген пікірталасқа қатыс \nБағалау: әрқайсыңа 1 қошқардан  \n\n\nЗадание 12. \nДискуссия \nНадеемся ты проснулся! Пора продолжить мини-исследование \nУчаствуй в дискуссии «Как технологии могут спасти мир будущего?» \nОценка по 1 кошкарчику каждому", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task13)


@bot.message_handler(content_types=["text"])
def event_Task13(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"T{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Дополнительно тезисы:Хотя IT может решить многие проблемы, важно учитывать этические проблемы, такие как конфиденциальность данных, цифровая безопасность и возможность усугубления существующего неравенства. Какие угрозы мы видим из дня сегодняшнего?", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task14)
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "13 тапсырма. \nЭкскурсия  \n Шағын зерттеу пішіні қалыптасып келеді, бәрекелді!  \nЭкскурсия кезінде барлық мектептерде суретке түс.  Ең соңында, оқығың келетін мектептің суретін таңда.  Неге дәл сол жерде оқығың келетінің туралы  қысқаша сипаттап фотомен бөліс, чат-ботқа жібер.  \nБағалау: орындалған тапсырма үшін 1 қошқардан \n\n\nЗадание 13. \nЭкскурсия \nМини-исследование обретает форму, так держать!\nВо время экскурсии сделай фото во всех школах. В самом конце, выбери фото со школой, где бы ты хотел учиться. И поделись фото с кратким описанием почему именно там ты бы хотел учиться, и отправьте в чат-бот.\nОценка по 1 кошкарчику за выполненное задание", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "13 тапсырма. \nЭкскурсия  \n Шағын зерттеу пішіні қалыптасып келеді, бәрекелді!  \nЭкскурсия кезінде барлық мектептерде суретке түс.  Ең соңында, оқығың келетін мектептің суретін таңда.  Неге дәл сол жерде оқығың келетінің туралы  қысқаша сипаттап фотомен бөліс, чат-ботқа жібер.  \nБағалау: орындалған тапсырма үшін 1 қошқардан \n\n\nЗадание 13. \nЭкскурсия \nМини-исследование обретает форму, так держать!\nВо время экскурсии сделай фото во всех школах. В самом конце, выбери фото со школой, где бы ты хотел учиться. И поделись фото с кратким описанием почему именно там ты бы хотел учиться, и отправьте в чат-бот.\nОценка по 1 кошкарчику за выполненное задание", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "13 тапсырма. \nЭкскурсия  \n Шағын зерттеу пішіні қалыптасып келеді, бәрекелді!  \nЭкскурсия кезінде барлық мектептерде суретке түс.  Ең соңында, оқығың келетін мектептің суретін таңда.  Неге дәл сол жерде оқығың келетінің туралы  қысқаша сипаттап фотомен бөліс, чат-ботқа жібер.  \nБағалау: орындалған тапсырма үшін 1 қошқардан \n\n\nЗадание 13. \nЭкскурсия \nМини-исследование обретает форму, так держать!\nВо время экскурсии сделай фото во всех школах. В самом конце, выбери фото со школой, где бы ты хотел учиться. И поделись фото с кратким описанием почему именно там ты бы хотел учиться, и отправьте в чат-бот.\nОценка по 1 кошкарчику за выполненное задание", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task14)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 1 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 1 кошкарчик! ')
            bot.send_message(chat_id, "13 тапсырма. \nЭкскурсия  \n Шағын зерттеу пішіні қалыптасып келеді, бәрекелді!  \nЭкскурсия кезінде барлық мектептерде суретке түс.  Ең соңында, оқығың келетін мектептің суретін таңда.  Неге дәл сол жерде оқығың келетінің туралы  қысқаша сипаттап фотомен бөліс, чат-ботқа жібер.  \nБағалау: орындалған тапсырма үшін 1 қошқардан \n\n\nЗадание 13. \nЭкскурсия \nМини-исследование обретает форму, так держать!\nВо время экскурсии сделай фото во всех школах. В самом конце, выбери фото со школой, где бы ты хотел учиться. И поделись фото с кратким описанием почему именно там ты бы хотел учиться, и отправьте в чат-бот.\nОценка по 1 кошкарчику за выполненное задание", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task14)


@bot.message_handler(content_types=["text"])
def event_Task14(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"U{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Поговорим о профессиях в IT и разберемся во всех предлагаемых трендах. Спроси себя: Что понятно? Что не понятно? Какие есть вопросы и предложения? Что по итогам первого дня возьму с собой в будущее? Все инсайты запиши в дневнике  notion Твоя цель собрать интересную тебе информацию об университете.  Возможно помогут вопросы: Какая миссия у университета? Что такое меритократия? Какие преимущества обучения в этом университете? Хочу ли я учиться здесь?Оформи заметки в notion.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        bot.send_message(chat_id, "14 тапсырма. \n«Мансап мүмкіндіктері жобасы» :\nТопта жобаны дайындап, идеяларды ұсыну.  \nСізді қазір мансап үшін қызықтыратын 4-5 IT мамандықтары туралы айтып беріңіз.  Топта талқылау үшін 20 минут беріледі.    \nӘр топтың сөйлеуіне 3 минут уақыт беріледі.  \nБағалау әр топқа 10 қошқардан \n\n\nЗадание 14.\n «Проект карьерные возможности»: \nПодготовить в команде проект и провести питчинг идеи.\nРасскажите о  4-5 IT- профессиях, которые вам наиболее интересны сейчас для карьеры. Групповое обсуждение в командах- 20 минут. \nВыступление каждой команды – по 3 минуты\nОценка по 10 кошкарчиков команде ", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task15)

@bot.message_handler(content_types=["text"])
def event_Task15(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"V{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Использовать информацию из дискуссии «Как технологии могут спасти мир будущего?»Обосновать идею.Вопросы: Где вы видите сейчас карьерные возможности? Какие сферы для реализации? Какие это могут быть профессии? Как вы это видите и какой вклад можете внести?", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, event_Task16)
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "15 тапсырма.\n Воркшоп: UX research, от Usability lab \nБіз шағын зерттеуімізді жалғастырамыз.\nСұраққа жауап бер: «UX research-тің болашақтағы 5 мамандықтың қызметіндегі рөлі қандай?» . Спикерге сұрақтар қойып, талқылаңыздар, жауап Notion-да ресімделіп, чат-бот арқылы жіберілуі керек. \nБағалау: әрқайсыңа 1 қошқардан \n\n\nЗадание 15. \nВоркшоп: UX research, от Usability lab \nПродолжаем мини-исследование!\nОтветь на вопрос: “Какая роль UX research в деятельности 5 профессий будущего?”. \nЗадай вопросы спикеру, обсудите, ответ нужно оформить в Notion и отправить через чат-бот.\nОценка по 1 кошкарчику каждому  ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "15 тапсырма.\n Воркшоп: UX research, от Usability lab \nБіз шағын зерттеуімізді жалғастырамыз.\nСұраққа жауап бер:«UX research-тің болашақтағы 5 мамандықтың қызметіндегі рөлі қандай?» . Спикерге сұрақтар қойып, талқылаңыздар, жауап Notion-да ресімделіп, чат-бот арқылы жіберілуі керек. \nБағалау: әрқайсыңа 1 қошқардан \n\n\nЗадание 15. \nВоркшоп: UX research, от Usability lab \nПродолжаем мини-исследование!\nОтветь на вопрос: “Какая роль UX research в деятельности 5 профессий будущего?”. \nЗадай вопросы спикеру, обсудите, ответ нужно оформить в Notion и отправить через чат-бот.\nОценка по 1 кошкарчику каждому  ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "15 тапсырма.\n Воркшоп: UX research, от Usability lab \nБіз шағын зерттеуімізді жалғастырамыз.\nСұраққа жауап бер: «UX research-тің болашақтағы 5 мамандықтың қызметіндегі рөлі қандай?» . Спикерге сұрақтар қойып, талқылаңыздар, жауап Notion-да ресімделіп, чат-бот арқылы жіберілуі керек. \nБағалау: әрқайсыңа 1 қошқардан \n\n\nЗадание 15. \nВоркшоп: UX research, от Usability lab \nПродолжаем мини-исследование!\nОтветь на вопрос: “Какая роль UX research в деятельности 5 профессий будущего?”. \nЗадай вопросы спикеру, обсудите, ответ нужно оформить в Notion и отправить через чат-бот.\nОценка по 1 кошкарчику каждому  ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task16)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "15 тапсырма.\n Воркшоп: UX research, от Usability lab \nБіз шағын зерттеуімізді жалғастырамыз.\nСұраққа жауап бер: «UX research-тің болашақтағы 5 мамандықтың қызметіндегі рөлі қандай?» . Спикерге сұрақтар қойып, талқылаңыздар, жауап Notion-да ресімделіп, чат-бот арқылы жіберілуі керек. \nБағалау: әрқайсыңа 1 қошқардан \n\n\nЗадание 15. \nВоркшоп: UX research, от Usability lab \nПродолжаем мини-исследование!\nОтветь на вопрос: “Какая роль UX research в деятельности 5 профессий будущего?”. \nЗадай вопросы спикеру, обсудите, ответ нужно оформить в Notion и отправить через чат-бот.\nОценка по 1 кошкарчику каждому  ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, event_Task16)

@bot.message_handler(content_types=["text"])
def event_Task16(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"W{count}"].value = message.text
    message_admin['size'] = message.text
    
    chat_id = message.chat.id
    if message.text == 'Help':
        bot.send_message(chat_id, "Не забудь, что вопросы эксперту можно задавать через чат-бот. Запиши свой ответ в notion.", reply_markup=types.ReplyKeyboardRemove())
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.register_next_step_handler(message, send_admin)
    else:
        app_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        app_markup.add(types.KeyboardButton('Help'))
        if group == 'Айтыскерлер':
            groupA = groupA + 1
            print(groupA)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Задание 17. \nBeginIT Pitch Day «5 IT профессий, которые помогут спасти мир»  Каждой команде нужно подготовить презентацию «5 IT профессий, которые помогут спасти мир» с учетом результатов мини-исследований, навыков и знаний, полученных во время воркшопов. Групповое обсуждение в командах- 25 минут. Питч презентации - 3-5 минут.\nОценка: вы оцениваете презентации всех 4 команд, в т.ч. и свою. Можно заработать от 1 до 4 кошкарчиков. ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Қолөнершілер':
            groupB = groupB + 1
            print(groupB)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Задание 17. \nBeginIT Pitch Day «5 IT профессий, которые помогут спасти мир»  Каждой команде нужно подготовить презентацию «5 IT профессий, которые помогут спасти мир» с учетом результатов мини-исследований, навыков и знаний, полученных во время воркшопов. Групповое обсуждение в командах- 25 минут. Питч презентации - 3-5 минут.\nОценка: вы оцениваете презентации всех 4 команд, в т.ч. и свою. Можно заработать от 1 до 4 кошкарчиков. ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Саудагерлер':
            groupC = groupC + 1
            print(groupC)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Задание 17. \nBeginIT Pitch Day «5 IT профессий, которые помогут спасти мир»  Каждой команде нужно подготовить презентацию «5 IT профессий, которые помогут спасти мир» с учетом результатов мини-исследований, навыков и знаний, полученных во время воркшопов. Групповое обсуждение в командах- 25 минут. Питч презентации - 3-5 минут.\nОценка: вы оцениваете презентации всех 4 команд, в т.ч. и свою. Можно заработать от 1 до 4 кошкарчиков. ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, send_admin)
        elif group == 'Саяхатшылар':
            groupD = groupD + 1
            print(groupD)
            photo_url = 'https://firebasestorage.googleapis.com/v0/b/begin-it-ef685.appspot.com/o/Mask%20group.png?alt=media&token=680eeaaa-2108-4199-ae74-6aced09cde52'
            bot.send_photo(chat_id, photo_url, caption='Сіз дұрыс жауап бердіңіз, cол себепті, 7 қошқар алыңыз! \n\n\nВы правильно ответили, поэтому получаете 7 кошкарчик! ')
            bot.send_message(chat_id, "Задание 17. \nBeginIT Pitch Day «5 IT профессий, которые помогут спасти мир»  Каждой команде нужно подготовить презентацию «5 IT профессий, которые помогут спасти мир» с учетом результатов мини-исследований, навыков и знаний, полученных во время воркшопов. Групповое обсуждение в командах- 25 минут. Питч презентации - 3-5 минут.\nОценка: вы оцениваете презентации всех 4 команд, в т.ч. и свою. Можно заработать от 1 до 4 кошкарчиков. ", reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, send_admin)




def send_admin(message):
    global count, groupA, groupB, groupC, groupD, group 
    worksheet[f"M{count}"].value = message.text
    count += 1
    workbook.save("ChatBot.xlsx")
    message_admin['partner'] = message.text
    first_name = message.chat.first_name
    chat_id = message.chat.id
    user_name = message.chat.username
    app_name, app_username = [], []
    app_name.append(first_name)
    app_username.append(user_name)
        
    app_name.clear()
    app_username.clear()
    
    bot.send_message(chat_id, "Поздравляю с успешным прохождением BeginIT", reply_markup=types.ReplyKeyboardRemove())

while True:
   bot.polling(none_stop=True)
   schedule.run_pending()
   time.sleep(1)
    
